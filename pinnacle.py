from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
import os

# excel manager

excelFile = 'Data.xlsx'
wb = Workbook()
wb = load_workbook(filename=excelFile)
sheetNumber = '1'
    
try:
    ws = wb.get_sheet_by_name(sheetNumber)

except:
    ws = wb.create_sheet(sheetNumber)
    
ws.append(['Pinnacle', '/////////', '/////////', '/////////', '/////////', '/////////'])

if (os.path.exists(excelFile)):
    os.remove(excelFile)
    wb.save(excelFile)
wb.save(excelFile)
    

#####################################

# setting up headless option

url = 'https://www.pinnacle.com/pt/tennis/matchups'

user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36 Edg/92.0.902.67'

options = webdriver.ChromeOptions()
# options.headless = True
options.add_argument(f'user-agent={user_agent}')
options.add_argument("--window-size=1920,1080")
options.add_argument('--ignore-certificate-errors')
options.add_argument('--allow-running-insecure-content')
options.add_argument("--disable-extensions")
options.add_argument("--proxy-server='direct://'")
options.add_argument("--proxy-bypass-list=*")
options.add_argument('--disable-gpu')
options.add_argument('--disable-dev-shm-usage')
options.add_argument('--no-sandbox')
# browser = webdriver.Chrome(executable_path="chromedriver.exe", options=options)
browser = webdriver.Chrome('chromedriver.exe')

time.sleep(1)

browser.maximize_window()

time.sleep(1)

browser.get(url)

time.sleep(5)

# tenis

parents_selector = 'div.style_row__3q4g_.style_row__3hCMX'

#print(len(parents_list))

event_list = []

id_list = []

while(True):
    
    time.sleep(2)
    
    parents_list = WebDriverWait(browser, 10).until(
        EC.presence_of_all_elements_located(
            (By.CSS_SELECTOR, parents_selector))
    )
    
    print(len(parents_list))
    
    try:
    
        if(parents_list[len(parents_list) - 1].id == id_list[len(id_list) - 1]):
            break
    except:
        pass
    
    
    for i in range(len(parents_list)):
    
        # scroll down page until element is visible
        #browser.execute_script('arguments[0].scrollIntoView();', parent)
        #print(parent.id)
        #print(parent.get_attribute('innerText'))
        
        
        
        should_continue = True
        
        for ids in id_list:
            if(parents_list[i].id == ids):
                should_continue = False
        
        if(should_continue == True):
                
            
            id_list.append(parents_list[i].id)
            
            childList = parents_list[i].find_elements(By.XPATH, '*')
            
            names = childList[0].get_attribute('innerText').split('\n')
            
            hrefElement = childList[0].find_element(By.XPATH, '*').find_element(By.XPATH, '*')
            href = hrefElement.get_attribute('href')
            
            
            try:
                
                odds = childList[2].get_attribute('innerText').split('\n')
                
                name1 = names[0]
                name2 = names[1]
                back1 = odds[0]
                lay1 = None
                back2 = odds[1]
                lay2 = None
                
                data = [name1, back1, lay1, 'x', name2, back2, lay2, href]
                event_list.append(data)
                
                ######
                #print(data)
                #time.sleep(1)
                #######
                
                #print(data)
                
                ws.append(data)
                
                if (os.path.exists(excelFile)):
                    os.remove(excelFile)
                    wb.save(excelFile)
                wb.save(excelFile)
            
            except:
                pass
            
        if(i == len(parents_list) - 1):
                
            browser.execute_script('arguments[0].scrollIntoView();', parents_list[i])
            
    
time.sleep(5)
    
if (os.path.exists(excelFile)):
    os.remove(excelFile)
    wb.save(excelFile)
wb.save(excelFile)

#print(event_list)

browser.close()

#
