from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from openpyxl import Workbook
from openpyxl import load_workbook
import os

# excel manager

excelFile = 'Data.xlsx'
wb = Workbook()
wb = load_workbook(filename=excelFile)
sheetNumber = '1'

# # create new sheets inside the excel file
# while (True):

#     try:
#         ws = wb.get_sheet_by_name(sheetNumber)

#     except:
#         ws = wb.create_sheet(sheetNumber)
#         wb.save(excelFile)
#         break

#     sheetNumber = str(int(sheetNumber) + 1)
    
try:
    ws = wb.get_sheet_by_name(sheetNumber)

except:
    ws = wb.create_sheet(sheetNumber)
    
ws.append(['Bet Fair', '/////////', '/////////', '/////////', '/////////', '/////////'])

if (os.path.exists(excelFile)):
    os.remove(excelFile)
    wb.save(excelFile)
wb.save(excelFile)

#####################################

# setting up headless option

url = 'https://www.betfair.com/exchange/plus/pt/t%C3%AAnis-apostas-2/today'

user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36 Edg/92.0.902.67'

options = webdriver.ChromeOptions()
# options.headless = True
options.add_argument(f'user-agent={user_agent}')
options.add_argument("--window-size=1920,1080")
options.add_argument('--ignore-certificate-errors')
options.add_argument('--allow-running-insecure-content')
options.add_argument("--disable-extensions")
options.add_argument("--proxy-server='direct://'")
options.add_argument("--proxy-bypass-list=*")
options.add_argument('--disable-gpu')
options.add_argument('--disable-dev-shm-usage')
options.add_argument('--no-sandbox')
# browser = webdriver.Chrome(executable_path="chromedriver.exe", options=options)
browser = webdriver.Chrome('chromedriver.exe')

time.sleep(1)

browser.maximize_window()

time.sleep(1)

browser.get(url)

# change filter to date

# filter_xpath = '/html/body/ui-view/div/div/div[2]/div/ui-view/ui-view/div/div/div/div/div[1]/div/div[1]/bf-super-coupon/main/ng-include[2]/section/div/div/bf-select'


# filter_selector = WebDriverWait(browser, 10).until(
#     EC.presence_of_element_located(
#         (By.XPATH, filter_xpath))
# )

# print(filter_selector.get_attribute('innerText'))

# dropdown = Select(filter_selector)
# dropdown.select_by_index(1)

######################

#time.sleep(5)

# tenis

list_selector = 'td.coupon-runners'
players_selector = 'ul.runners'
next_page_xpath = '/html/body/ui-view/div/div/div[2]/div/ui-view/ui-view/div/div/div/div/div[1]/div/div[1]/bf-super-coupon/main/ng-include[4]/bf-coupon-page-navigation/ul/li[4]/a'

event_list = []

# getting players and odds

current_page = 1

page_url = url

while(True):
    
    try:
        
        all_list = WebDriverWait(browser, 10).until(
            EC.presence_of_all_elements_located(
                (By.CSS_SELECTOR, list_selector))
        )

        players_list = WebDriverWait(browser, 10).until(
            EC.presence_of_all_elements_located(
                (By.CSS_SELECTOR, players_selector))
        )

        # next_page = WebDriverWait(browser, 10).until(
        #     EC.presence_of_element_located(
        #         (By.XPATH, next_page_xpath))
        # )
        
        print(browser.current_url)
        
        if(browser.current_url != page_url):
            break
        
        i = 0

        print(len(all_list))
        print(len(players_list))

    
        for all in all_list:
        
            #print('//////////////////////////////')
        
            browser.execute_script('arguments[0].scrollIntoView();', all)
            # print(all.get_attribute('innerText'))
            # print(all.get_attribute('href'))
            
            ##### getting href
            parent = all.find_element(By.XPATH, '..')
            href_child = parent.find_element(By.XPATH, '*')
            href_child = href_child.find_element(By.XPATH, '*')
            #####
            
            ##### getting players
            players = players_list[i].get_attribute('innerText').split('\n')
            #####
            
            info = all.get_attribute('innerText').split('\n')
            
            name1 = ''
            name2 = ''
            back1 = ''
            lay1 = ''
            back2 = ''
            lay2 = ''
            href = ''
            
            try:
            
                name1 = players[0]
                name2 = players[1]
                back1 = info[0]
                lay1 = info[2]
                back2 = info[4]
                lay2 = info[6]
                href = href_child.get_attribute('href')
            
            except:
                pass
                
            data = [name1, back1, lay1, 'x', name2, back2, lay2, href]
            
            event_list.append(data)
            
            ws.append(data)
            
            if (os.path.exists(excelFile)):
                os.remove(excelFile)
                wb.save(excelFile)
            wb.save(excelFile)
            
            #print(data)
            
            i+=1
        
        #next_page.click()
        
        current_page += 1
        page_url = url + '/' + str(current_page)
        browser.get(page_url)
        
        
    
    except:
        break

ws.append(['/////////', '/////////', '/////////', '/////////', '/////////'])

if (os.path.exists(excelFile)):
    os.remove(excelFile)
    wb.save(excelFile)
wb.save(excelFile)

browser.close()

#
