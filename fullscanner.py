print('Loading Betfair...')
import betfair
print('Loading S Markets...')
import smarkets
print('Loading Pinnacle...')
import pinnacle

names_betfair = []
names_smarkets = []
names_pinnacle = []

for i in range(len(betfair.event_list)):
    names_betfair.append(betfair.event_list[i][0])
    
for i in range(len(smarkets.event_list)):
    names_smarkets.append(smarkets.event_list[i][0])
    
for i in range(len(pinnacle.event_list)):
    names_pinnacle.append(pinnacle.event_list[i][0])

# indexs_betfair = []
# indexs_smarkets = []
# indexs_pinnacle = []

# i = 0
# j = 0

# for name1 in names1:
    
#     j = 0
    
#     for name2 in names2:
#         first_name1 = name1.split(' ')
#         first_name2 = name2.split(' ')
        
#         if(first_name1 == first_name2):
#             indexs_betfair.append(i)
#             indexs_smarkets.append(j)
        
#         j += 1
    
#     i += 1

# opportunities_betfair = []
# opportunities_smarkets = []

# for i in indexs_betfair:
#     print(betfair.event_list[i])
#     opportunities_betfair.append(betfair.event_list[i])
    
# for i in indexs_smarkets:
#     print(smarkets.event_list[i])
#     opportunities_smarkets.append(smarkets.event_list[i])

i = 0
j = 0

opportunities = []
intersection_games = []

#comparing betfair x smarkets

for name_betfair in names_betfair:
    
    j = 0
    
    for name_smarkets in names_smarkets:
        first_name_betfair = name_betfair.split(' ')
        first_name_smarkets = name_smarkets.split(' ')
        
        if(first_name_betfair == first_name_smarkets):
            
            back1_betfair = float(betfair.event_list[i][1].replace(',', '.'))
            back2_betfair = float(betfair.event_list[i][5].replace(',', '.'))
            back1_smarkets = float(smarkets.event_list[j][1].replace(',', '.'))
            back2_smarkets = float(smarkets.event_list[j][5].replace(',', '.'))
  
            
            intersection_games.append([
                betfair.event_list[i][0], betfair.event_list[i][4], back1_betfair, back2_betfair, betfair.event_list[i][7],  
                smarkets.event_list[j][0], smarkets.event_list[j][4], back1_smarkets, back2_smarkets, smarkets.event_list[j][7],   
                ])
            
            if(back1_betfair >= back1_smarkets):
                best_back1 = back1_betfair
                href1 = betfair.event_list[i][7]
            else:
                best_back1 = back1_smarkets
                href1 = smarkets.event_list[j][7]
            
            if(back2_betfair >= back2_smarkets):
                best_back2 = back2_betfair
                href2 = betfair.event_list[i][7]
            else:
                best_back2 = back2_smarkets
                href2 = smarkets.event_list[j][7]
                
            expect_value = (1/best_back1) + (1/best_back2)
            
            opportunities.append([best_back1, best_back2, expect_value, href1, href2])
        
        j += 1
    
    i += 1
    
#comparing betfair x pinnacle

i = 0
j = 0
    
for name_betfair in names_betfair:
    
    j = 0
    
    for name_pinnacle in names_pinnacle:
        first_name_betfair = name_betfair.split(' ')
        first_name_pinnacle = name_pinnacle.split(' ')
        
        if(first_name_betfair == first_name_pinnacle):
            
            back1_betfair = float(betfair.event_list[i][1].replace(',', '.'))
            back2_betfair = float(betfair.event_list[i][5].replace(',', '.'))
            back1_pinnacle = float(pinnacle.event_list[j][1].replace(',', '.'))
            back2_pinnacle = float(pinnacle.event_list[j][5].replace(',', '.'))
            
            intersection_games.append([
                betfair.event_list[i][0], betfair.event_list[i][4], back1_betfair, back2_betfair, betfair.event_list[i][7],  
                pinnacle.event_list[j][0], pinnacle.event_list[j][4], back1_pinnacle, back2_pinnacle, pinnacle.event_list[j][7],   
                ])
            
            if(back1_betfair >= back1_pinnacle):
                best_back1 = back1_betfair
                href1 = betfair.event_list[i][7]
            else:
                best_back1 = back1_pinnacle
                href1 = pinnacle.event_list[j][7]
            
            if(back2_betfair >= back2_pinnacle):
                best_back2 = back2_betfair
                href2 = betfair.event_list[i][7]
            else:
                best_back2 = back2_pinnacle
                href2 = pinnacle.event_list[j][7]
                
            expect_value = (1/best_back1) + (1/best_back2)
            
            opportunities.append([best_back1, best_back2, expect_value, href1, href2])
            
        
        j += 1
    
    i += 1
    
#comparing smarkets x pinnacle

i = 0
j = 0

for name_smarkets in names_smarkets:
    
    j = 0
    
    for name_pinnacle in names_pinnacle:
        first_name_smarkets = name_smarkets.split(' ')
        first_name_pinnacle = name_pinnacle.split(' ')
        
        if(first_name_smarkets == first_name_pinnacle):
            
            back1_smarkets = float(smarkets.event_list[i][1].replace(',', '.'))
            back2_smarkets = float(smarkets.event_list[i][5].replace(',', '.'))
            back1_pinnacle = float(pinnacle.event_list[j][1].replace(',', '.'))
            back2_pinnacle = float(pinnacle.event_list[j][5].replace(',', '.'))
            
            intersection_games.append([
                smarkets.event_list[i][0], smarkets.event_list[i][4], back1_smarkets, back2_smarkets, smarkets.event_list[i][7],  
                pinnacle.event_list[j][0], pinnacle.event_list[j][4], back1_pinnacle, back2_pinnacle, pinnacle.event_list[j][7],   
                ])
            
            if(back1_smarkets >= back1_pinnacle):
                best_back1 = back1_smarkets
                href1 = smarkets.event_list[i][7]
            else:
                best_back1 = back1_pinnacle
                href1 = pinnacle.event_list[j][7]
            
            if(back2_smarkets >= back2_pinnacle):
                best_back2 = back2_smarkets
                href2 = smarkets.event_list[i][7]
            else:
                best_back2 = back2_pinnacle
                href2 = pinnacle.event_list[j][7]
            
            expect_value = (1/best_back1) + (1/best_back2)
            
            opportunities.append([best_back1, best_back2, expect_value, href1, href2])
            
            
        
        j += 1
    
    i += 1
    
from openpyxl import Workbook
from openpyxl import load_workbook
import os


# excel manager

excelFile = 'Oportunities.xlsx'
wb = Workbook()
wb = load_workbook(filename=excelFile)
sheetNumber = 'Plan1'
ws = wb.get_sheet_by_name(sheetNumber)

from datetime import datetime

now = datetime.now()
current_time = now.strftime('%H:%M:%S')
    
# ws.append([current_time])
# ws.append(['BETFAIR', '//////////////', '//////////////', '//////////////', '//////////////'])
# for op in opportunities_betfair:
#     ws.append(op)

# ws.append(['SMARKETS', '//////////////', '//////////////', '//////////////', '//////////////'])
# for op in opportunities_smarkets:
#     ws.append(op)
0
print(opportunities)
#print(intersection_games)


for op in opportunities:
    ws.append(op)

if (os.path.exists(excelFile)):
    os.remove(excelFile)
    wb.save(excelFile)
wb.save(excelFile)
    
