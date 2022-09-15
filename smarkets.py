from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
import os

# excel manager

excelFile = 'Data.xlsx'
wb = Workbook()
wb = load_workbook(filename=excelFile)
sheetNumber = '1'

# # create new sheets inside the excel file
# while (True):

#     try:
#         ws = wb.get_sheet_by_name(sheetNumber)

#     except:
#         ws = wb.create_sheet(sheetNumber)
#         wb.save(excelFile)
#         break

#     sheetNumber = str(int(sheetNumber) + 1)
    
try:
    ws = wb.get_sheet_by_name(sheetNumber)

except:
    ws = wb.create_sheet(sheetNumber)
    
ws.append(['S Markets', '/////////', '/////////', '/////////', '/////////', '/////////'])

if (os.path.exists(excelFile)):
    os.remove(excelFile)
    wb.save(excelFile)
wb.save(excelFile)
    

#####################################

# setting up headless option

url = 'https://smarkets.com/listing/sport/tennis?period=today'

user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36 Edg/92.0.902.67'

options = webdriver.ChromeOptions()
# options.headless = True
options.add_argument(f'user-agent={user_agent}')
options.add_argument("--window-size=1920,1080")
options.add_argument('--ignore-certificate-errors')
options.add_argument('--allow-running-insecure-content')
options.add_argument("--disable-extensions")
options.add_argument("--proxy-server='direct://'")
options.add_argument("--proxy-bypass-list=*")
options.add_argument('--disable-gpu')
options.add_argument('--disable-dev-shm-usage')
options.add_argument('--no-sandbox')
# browser = webdriver.Chrome(executable_path="chromedriver.exe", options=options)
browser = webdriver.Chrome('chromedriver.exe')

time.sleep(1)

browser.maximize_window()

time.sleep(1)

browser.get(url)


time.sleep(5)

# tenis

#list_selector = 'div.contract-items.row.open'
list_selector = 'li.item-tile.event-tile.upcoming.layout-row'

# players_list = WebDriverWait(browser, 10).until(
#     EC.presence_of_all_elements_located(
#         (By.CSS_SELECTOR, players_list_selector))
# )

# odds_list = WebDriverWait(browser, 10).until(
#     EC.presence_of_all_elements_located(
#         (By.CSS_SELECTOR, odds_list_selector))
# )

all_list = WebDriverWait(browser, 10).until(
    EC.presence_of_all_elements_located(
        (By.CSS_SELECTOR, list_selector))
)

i = 0

print(len(all_list))

event_list = []

# getting players and odds
for all in all_list:

    # 0 - name1
    # 1 - back1
    # 3 - lay1
    # 5 - name2
    # 6 - back2
    # 8 - lay2

    # scroll down page until element is visible
    browser.execute_script('arguments[0].scrollIntoView();', all)
        
    try:
        childList = all.find_elements(By.XPATH, '*')
        
        hrefElement = childList[0].find_element(By.XPATH, '*')
        
        info = childList[1].get_attribute('innerText').split('\n')
        
 
        #print(info)

        name1 = info[0]
        back1 = info[1]
        lay1 = info[3]
        name2 = info[5]
        back2 = info[6]
        lay2 = info[8]
        href = hrefElement.get_attribute('href')

        if (name1 != 'Yes' and name2 != 'No' and back1 != 'LICITAÇÃO' and back2!= 'LICITAÇÃO' and lay1 != 'LICITAÇÃO' and lay2 != 'LICITAÇÃO'):

            data = [name1, back1, lay1, 'x', name2, back2, lay2, href]
            event_list.append(data)
            #print(data)
            ws.append(data)
            
            if (os.path.exists(excelFile)):
                os.remove(excelFile)
                wb.save(excelFile)
            wb.save(excelFile)

    except:
        pass

    # time.sleep(1)
    
ws.append(['/////////', '/////////', '/////////', '/////////', '/////////'])

if (os.path.exists(excelFile)):
    os.remove(excelFile)
    wb.save(excelFile)
wb.save(excelFile)

#print(event_list)

browser.close()

#
